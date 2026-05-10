import streamlit as st
from supabase import create_client, Client
from datetime import datetime
import pandas as pd
import time
import urllib.parse
import io
import os
import tempfile
import urllib.request
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Używamy fpdf do prawdziwych plików PDF
try:
    from fpdf import FPDF
except ImportError:
    st.error("Błąd: Brak biblioteki fpdf. Wpisz w terminalu: pip install fpdf2")

# --- KONFIGURACJA ---
URL = "https://hdmptdcuqxqutfgrgmrj.supabase.co"
KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImhkbXB0ZGN1cXhxdXRmZ3JnbXJqIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY3NzQ2NTksImV4cCI6MjA5MjM1MDY1OX0.ZI18vTCpYloVOdzpZuVHYVH2OwKJMsrQINgaJNl-vho"

supabase: Client = create_client(URL, KEY)

st.set_page_config(page_title="Zamówienia Pakamera", page_icon="🛒", layout="wide")

# --- STAN SESJI ---
if 'zalogowany' not in st.session_state:
    st.session_state.zalogowany = False
    st.session_state.uzytkownik = ""
    st.session_state.rola = "użytkownik"

# --- LOGOWANIE ---
if not st.session_state.zalogowany:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.title("🛒 ZAMÓWIENIA")
        st.caption("System zarządzania materiałami")
        with st.container(border=True):
            l = st.text_input("Login")
            p = st.text_input("Hasło", type="password")
            if st.button("ZALOGUJ", use_container_width=True, type="primary"):
                if l == "Emil" and p == "Sosna100%":
                    st.session_state.zalogowany = True
                    st.session_state.uzytkownik = "Emil"
                    st.session_state.rola = "admin"
                    st.rerun()
                else:
                    res = supabase.table("pracownicy").select("*").eq("login", l).execute()
                    if res.data:
                        user_data = res.data[0]
                        db_password = user_data.get('hasło') or user_data.get('haslo')
                        if db_password == p:
                            st.session_state.zalogowany = True
                            st.session_state.uzytkownik = l
                            st.session_state.rola = user_data.get('rola') or "użytkownik"
                            st.rerun()
                        else: st.error("Błędne hasło!")
                    else: st.error("Użytkownik nie istnieje!")
else:
    # --- MENU BOCZNE ---
    with st.sidebar:
        st.success(f"Zalogowano: **{st.session_state.uzytkownik}**")
        st.divider()
        opcje = ["📝 Nowe Zamówienie", "📋 Moje Aktywne", "🔎 Historia i Raporty", "📊 Statystyki", "💬 Czat / Sugestie", "📖 Instrukcja"]
        if st.session_state.rola == "admin":
            opcje.insert(1, "⚙️ Panel Realizacji (Admin)")
            opcje.insert(4, "👥 Zarządzanie Kontami")
        menu = st.radio("MENU", opcje)
        st.divider()
        if st.button("🔄 Odśwież dane", use_container_width=True): st.rerun()
        if st.button("🚪 Wyloguj", use_container_width=True):
            st.session_state.zalogowany = False
            st.rerun()

    status_emoji = {"Oczekujące": "⏳", "Zamówione": "🚚", "Niedostępne": "❌", "Zamiennik": "🔄", "Zrealizowane": "✅"}

    def render_status_alert(status_text):
        ikona = status_emoji.get(status_text, "🔹")
        msg = f"**Status:** {ikona} {status_text}"
        if status_text == "Zrealizowane": st.success(msg)
        elif status_text in ["Niedostępne", "Zamiennik"]: st.error(msg)
        elif status_text == "Oczekujące": st.warning(msg)
        else: st.info(msg)

    # --- FUNKCJA PDF ---
    def make_real_pdf(df, title_text):
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.add_page()
        def cln(text):
            if pd.isna(text) or text is None: return "-"
            text = str(text)
            mapping = {'ą':'a','ć':'c','ę':'e','ł':'l','ń':'n','ó':'o','ś':'s','ź':'z','ż':'z','Ą':'A','Ć':'C','Ę':'E','Ł':'L','Ń':'N','Ó':'O','Ś':'S','Ź':'Z','Ż':'Z'}
            for k, v in mapping.items(): text = text.replace(k, v)
            return text.encode('ascii', 'ignore').decode('ascii')
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, cln(title_text), ln=1, align="C")
        pdf.ln(3)
        pdf.set_font("Arial", "", 5.5)
        cols = ["Zglosz.", "Realiz.", "Pozycja", "Wymiary", "Material", "Ilosc", "Projekt", "Pilnosc", "Zglosil", "Status", "Uwagi"]
        widths = [14, 14, 26, 14, 14, 8, 16, 12, 16, 16, 38]
        pdf.set_fill_color(47, 117, 181)
        pdf.set_text_color(255, 255, 255)
        for i, col in enumerate(cols): pdf.cell(widths[i], 6, cln(col), border=1, align="C", fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        for _, row in df.iterrows():
            pdf.cell(widths[0], 5, cln(row['Zgłoszono'])[:10], border=1)
            pdf.cell(widths[1], 5, cln(row['Zrealizowano'])[:10], border=1)
            pdf.cell(widths[2], 5, cln(row['Pozycja'])[:20], border=1)
            pdf.cell(widths[3], 5, cln(row['Wymiary'])[:10], border=1)
            pdf.cell(widths[4], 5, cln(row['Materiał'])[:10], border=1)
            pdf.cell(widths[5], 5, cln(row['Ilość'])[:6], border=1)
            pdf.cell(widths[6], 5, cln(row['Projekt'])[:12], border=1)
            pdf.cell(widths[7], 5, cln(row['Pilność'])[:8], border=1)
            pdf.cell(widths[8], 5, cln(row['Zgłaszający'])[:12], border=1)
            pdf.cell(widths[9], 5, cln(row['Status'])[:12], border=1)
            pdf.cell(widths[10], 5, cln(row['Uwagi'])[:32], border=1)
            pdf.ln()
        try:
            out = pdf.output(dest='S')
            if isinstance(out, str): return out.encode('latin1')
            return bytes(out)
        except TypeError: return bytes(pdf.output())

    # =========================================================================
    # ZAKŁADKI
    # =========================================================================
    if menu == "📝 Nowe Zamówienie":
        st.title("📝 Dodaj zamówienie")
        admins_res = supabase.table("pracownicy").select("login, telefon").eq("rola", "admin").execute()
        admin_phones = {a['login']: a['telefon'] for a in admins_res.data if a.get('telefon')}
        
        with st.container(border=True):
            pozycja = st.text_input("🔧 Pozycja (np. Śruba zamkowa)")
            col1, col2 = st.columns(2)
            wymiary = col1.text_input("📏 Wymiary (np. M8x40)")
            material = col2.text_input("🧱 Materiał (np. Ocynk)")
            col3, col4 = st.columns(2)
            ilosc = st.text_input("🔢 Ilość (np. 100 szt.)")
            pilnosc = col4.selectbox("🚨 Pilność", ["Normalna", "PILNE ⚡", "KRYTYCZNE 🛑"])
            projekt = st.text_input("🏗️ Projekt / Cel")

            st.divider()
            st.subheader("📸 Załącznik (Zdjęcie lub PDF)")
            zal_col1, zal_col2 = st.columns(2)
            zdjecie_cam = None
            if zal_col1.toggle("📷 Użyj aparatu"): zdjecie_cam = st.camera_input("Zrób zdjęcie")
            plik_upload = zal_col2.file_uploader("📁 Wybierz plik", type=["jpg", "jpeg", "png", "pdf"])

            st.divider()
            opcje_wa = ["-- Nie wysyłaj --"] + list(admin_phones.keys())
            powiadom = st.selectbox("📲 Powiadom admina (WhatsApp):", opcje_wa)
            
            if st.button("WYŚLIJ ZAMÓWIENIE", type="primary", use_container_width=True):
                if pozycja and ilosc:
                    url_zdj = ""
                    final_file = None
                    ext = "jpg"
                    if zdjecie_cam: final_file = zdjecie_cam.getvalue()
                    elif plik_upload:
                        final_file = plik_upload.getvalue()
                        ext = plik_upload.name.split('.')[-1]

                    if final_file:
                        nazwa = f"{int(time.time())}_{st.session_state.uzytkownik}.{ext}"
                        c_type = "application/pdf" if ext.lower() == "pdf" else "image/jpeg"
                        supabase.storage.from_("zdjecia_zamowien").upload(nazwa, final_file, {"content-type": c_type})
                        url_zdj = supabase.storage.from_("zdjecia_zamowien").get_public_url(nazwa)

                    supabase.table("zamowienia").insert({
                        "pozycja": pozycja, "wymiary": wymiary, "material": material, "ilosc": ilosc, 
                        "projekt": projekt, "pilnosc": pilnosc, "status": "Oczekujące", 
                        "zgloszone_przez": st.session_state.uzytkownik, "data_zgloszenia": str(datetime.today().date()),
                        "zdjecie_url": url_zdj
                    }).execute()
                    
                    st.success("Wysłano pomyślnie!")
                    if powiadom != "-- Nie wysyłaj --":
                        nr = "".join(c for c in admin_phones[powiadom] if c.isdigit())
                        t = f"Nowe zamówienie: {pozycja}. Od: {st.session_state.uzytkownik}"
                        st.link_button("📲 Wyślij WhatsApp", f"https://wa.me/{nr}?text={urllib.parse.quote(t)}", use_container_width=True)
                    else:
                        time.sleep(1.5); st.rerun()
                else:
                    st.error("Wypełnij wymagane pola!")

    elif menu == "⚙️ Panel Realizacji (Admin)":
        st.title("⚙️ Zarządzanie realizacją")
        prac_res = supabase.table("pracownicy").select("login, telefon").execute()
        pracownicy_dict = {p['login']: p.get('telefon', '') for p in prac_res.data}
        res = supabase.table("zamowienia").select("*").neq("status", "Zrealizowane").order("id", desc=True).execute()
        
        if not res.data: st.success("Wszystkie zamówienia zrealizowane! 👏")
        else:
            for r in res.data:
                with st.container(border=True):
                    st.markdown(f"### 📦 {r['pozycja'].upper()} `x {r['ilosc']}`")
                    st.markdown(f"👤 `{r['zgloszone_przez']}` | 🏗️ `{r.get('projekt') or '-'}` | 🚨 `{r.get('pilnosc') or 'Normalna'}` | 📏 `{r.get('wymiary') or '-'}`")
                    
                    c_st, c_uw, c_zap = st.columns([2, 4, 1])
                    st_list = ["Oczekujące", "Zamówione", "Niedostępne", "Zamiennik", "Zrealizowane"]
                    n_st = c_st.selectbox("Status", st_list, index=st_list.index(r['status']), key=f"st_sel_{r['id']}", label_visibility="collapsed")
                    n_uw = c_uw.text_input("Notatka", value=r.get('uwagi_admina') or "", key=f"uw_inp_{r['id']}", placeholder="Dodaj notatkę...", label_visibility="collapsed")
                    
                    if c_zap.button("💾 Zapisz", key=f"save_{r['id']}", type="primary", use_container_width=True):
                        u_data = {"status": n_st, "uwagi_admina": n_uw}
                        if n_st == "Zrealizowane": u_data["data_realizacji"] = str(datetime.today().date())
                        else: u_data["data_realizacji"] = None
                        supabase.table("zamowienia").update(u_data).eq("id", r['id']).execute()
                        st.toast("Zapisano!"); time.sleep(0.5); st.rerun()

                    b1, b2, b3, b4 = st.columns([2, 2, 1, 1])
                    msg = f"Aktualizacja: {r['pozycja']} | Status: {n_st} | Uwagi: {n_uw}"
                    tel_zgl = pracownicy_dict.get(r['zgloszone_przez'], '')
                    
                    if tel_zgl:
                        nr_zgl = "".join(filter(str.isdigit, tel_zgl))
                        b1.link_button(f"📲 WA: {r['zgloszone_przez']}", f"https://wa.me/{nr_zgl}?text={urllib.parse.quote(msg)}", use_container_width=True)
                    else:
                        b1.button(f"⚠️ Brak Tel", disabled=True, use_container_width=True)
                    
                    with b2.popover("➕ WA Inni"):
                        lista_opcji = [k for k in pracownicy_dict.keys() if k != r['zgloszone_przez']]
                        zaint = st.multiselect("Dodaj osoby:", lista_opcji, key=f"multi_{r['id']}")
                        for os in zaint:
                            t_os = pracownicy_dict.get(os, '')
                            if t_os:
                                n_os = "".join(filter(str.isdigit, t_os))
                                st.link_button(f"Wyślij do: {os}", f"https://wa.me/{n_os}?text={urllib.parse.quote(msg)}", use_container_width=True)

                    with b3.popover("⚙️ Akcje"):
                        up_poz = st.text_input("Pozycja", value=r['pozycja'], key=f"ep_{r['id']}")
                        up_ilo = st.text_input("Ilość", value=r['ilosc'], key=f"ei_{r['id']}")
                        up_wym = st.text_input("Wymiary", value=r.get('wymiary',''), key=f"ew_{r['id']}")
                        up_mat = st.text_input("Materiał", value=r.get('material',''), key=f"em_{r['id']}")
                        up_pro = st.text_input("Projekt", value=r.get('projekt',''), key=f"ej_{r['id']}")
                        up_pil = st.selectbox("Pilność", ["Normalna", "PILNE ⚡", "KRYTYCZNE 🛑"], index=["Normalna", "PILNE ⚡", "KRYTYCZNE 🛑"].index(r.get('pilnosc','Normalna')), key=f"epi_{r['id']}")
                        if st.button("Zapisz zmiany", key=f"btn_edit_{r['id']}", type="primary", use_container_width=True):
                            supabase.table("zamowienia").update({"pozycja": up_poz, "ilosc": up_ilo, "wymiary": up_wym, "material": up_mat, "projekt": up_pro, "pilnosc": up_pil}).eq("id", r['id']).execute()
                            st.rerun()
                        st.divider()
                        if st.button("🗑️ Usuń trwale", key=f"del_{r['id']}", use_container_width=True):
                            supabase.table("zamowienia").delete().eq("id", r['id']).execute(); st.rerun()

                    if r.get('zdjecie_url'):
                        if r['zdjecie_url'].lower().endswith(".pdf"): b4.link_button("📄 PDF", r['zdjecie_url'], use_container_width=True)
                        else:
                            with b4.popover("🖼️ FOTO"): st.image(r['zdjecie_url'], use_container_width=True)
                    else:
                        b4.button("❌ 🖼️", disabled=True, use_container_width=True)

    elif menu == "📋 Moje Aktywne":
        st.title("📋 Twoje zamówienia")
        res = supabase.table("zamowienia").select("*").eq("zgloszone_przez", st.session_state.uzytkownik).neq("status", "Zrealizowane").execute()
        if not res.data: st.info("Brak aktywnych zamówień.")
        for r in res.data:
            with st.container(border=True):
                st.markdown(f"### 📦 {r['pozycja'].upper()} `x {r['ilosc']}`")
                render_status_alert(r['status'])
                if r.get('uwagi_admina'): st.info(f"Odpis Admina: {r['uwagi_admina']}")

    elif menu == "🔎 Historia i Raporty":
        st.title("🔎 Pełna baza zamówień")
        
        res_f = supabase.table("zamowienia").select("projekt, zgloszone_przez").execute()
        projekty = sorted(list(set([x['projekt'] for x in res_f.data if x.get('projekt')])))
        osoby = sorted(list(set([x['zgloszone_przez'] for x in res_f.data if x.get('zgloszone_przez')])))
        
        with st.container(border=True):
            f_col1, f_col2, f_col3, f_col4 = st.columns(4)
            f_slowo = f_col1.text_input("🔍 Szukaj nazwy...")
            f_proj = f_col2.selectbox("🏗️ Projekt", ["-- Wszystkie --"] + projekty)
            f_kto = f_col3.selectbox("👤 Kto zgłosił", ["-- Wszyscy --"] + osoby)
            f_status = f_col4.selectbox("📌 Status", ["-- Wszystkie --", "Oczekujące", "Zamówione", "Niedostępne", "Zamiennik", "Zrealizowane"])
            
            q = supabase.table("zamowienia").select("*")
            if f_proj != "-- Wszystkie --": q = q.eq("projekt", f_proj)
            if f_kto != "-- Wszyscy --": q = q.eq("zgloszone_przez", f_kto)
            if f_status != "-- Wszystkie --": q = q.eq("status", f_status)
            
            wynik = q.order("id", desc=True).execute().data
            if f_slowo:
                wynik = [x for x in wynik if f_slowo.lower() in x['pozycja'].lower()]

        if wynik:
            df_h = pd.DataFrame(wynik)
            st.dataframe(df_h, use_container_width=True, hide_index=True)
            
            df_export = df_h[["data_zgloszenia", "data_realizacji", "pozycja", "wymiary", "material", "ilosc", "projekt", "pilnosc", "zgloszone_przez", "status", "uwagi_admina"]].copy()
            df_export.columns = ["Zgłoszono", "Zrealizowano", "Pozycja", "Wymiary", "Materiał", "Ilość", "Projekt", "Pilność", "Zgłaszający", "Status", "Uwagi"]
            
            min_date = df_export['Zgłoszono'].min() if not df_export.empty else "Brak"
            max_date = df_export['Zgłoszono'].max() if not df_export.empty else "Brak"
            zakres_dat = f"{min_date} do {max_date}"

            col_ex1, col_ex2 = st.columns(2)

            try:
                output_excel = io.BytesIO()
                with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                    df_export.to_excel(writer, index=False, sheet_name='Historia', startrow=2)
                    ws = writer.sheets['Historia']
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_export.columns))
                    cell_title = ws.cell(row=1, column=1)
                    cell_title.value = f"HISTORIA ZAMÓWIEŃ (Zakres: {zakres_dat})"
                    cell_title.font = Font(size=14, bold=True, color="FFFFFF")
                    cell_title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell_title.alignment = Alignment(horizontal="center", vertical="center")
                    
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    for col_num in range(1, len(df_export.columns) + 1):
                        c = ws.cell(row=3, column=col_num)
                        c.font = Font(bold=True, color="FFFFFF")
                        c.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        c.border = thin_border
                    
                    for row in ws.iter_rows(min_row=4, max_row=len(df_export)+3, min_col=1, max_col=len(df_export.columns)):
                        for c in row:
                            c.border = thin_border
                            c.alignment = Alignment(horizontal="left", vertical="center")
                    
                    for i, col in enumerate(df_export.columns, 1): ws.column_dimensions[get_column_letter(i)].width = 18
                    
                    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
                    ws.page_setup.paperSize = ws.PAPERSIZE_A4
                    ws.page_setup.fitToPage = True
                    ws.page_setup.fitToHeight = 0
                    ws.page_setup.fitToWidth = 1
                
                col_ex1.download_button("📊 Pobierz Excel na dysk", output_excel.getvalue(), "historia_zamowien.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            except Exception as e:
                col_ex1.error("Błąd generowania Excela. Zainstaluj openpyxl")

            try:
                pdf_bytes = make_real_pdf(df_export, f"HISTORIA ZAMOWIEN (Zakres: {zakres_dat})")
                col_ex2.download_button("📄 Pobierz PDF na dysk", data=pdf_bytes, file_name="historia_zamowien.pdf", mime="application/pdf", use_container_width=True)
            except Exception as e:
                col_ex2.error(f"Wystąpił błąd podczas generowania PDF: {e}")

        else:
            st.info("Nie znaleziono zamówień spełniających kryteria.")

    elif menu == "👥 Zarządzanie Kontami":
        st.title("👥 Zarządzanie pracownikami")
        with st.container(border=True):
            st.subheader("➕ Dodaj nowe konto")
            c1, c2, c3, c4 = st.columns(4)
            n_log = c1.text_input("Login")
            n_has = c2.text_input("Hasło") 
            n_rol = c3.selectbox("Rola", ["użytkownik", "admin"])
            n_tel = c4.text_input("Telefon (np. 48123456789)")
            if st.button("Utwórz konto", type="primary"):
                if n_log and n_has:
                    supabase.table("pracownicy").insert({"login": n_log, "haslo": n_has, "rola": n_rol, "telefon": n_tel}).execute()
                    st.success(f"Dodano: {n_log}"); time.sleep(1); st.rerun()
                else: st.error("Login i Hasło są obowiązkowe!")

        st.divider()
        res_p = supabase.table("pracownicy").select("*").order("login").execute()
        for p in res_p.data:
            if not p.get('login'): continue
            with st.container(border=True):
                col_info, col_edit, col_del = st.columns([4, 1, 1])
                aktualne_haslo = p.get('hasło') or p.get('haslo') or ""
                haslo_widoczne = aktualne_haslo if aktualne_haslo else "???"
                if p['login'].lower() == "emil" and st.session_state.uzytkownik.lower() != "emil": haslo_widoczne = "••••••••"
                col_info.markdown(f"👤 **{p['login']}** | 🔑 Hasło: `{haslo_widoczne}` | 🛠️ Rola: `{p.get('rola')}` | 📞 Tel: `{p.get('telefon','')}`")
                
                with col_edit.popover("✏️ Edytuj"):
                    e_has = st.text_input("Nowe hasło", value=aktualne_haslo, key=f"eh_{p['login']}")
                    e_tel = st.text_input("Nowy telefon", value=p.get('telefon') or "", key=f"et_{p['login']}")
                    e_rol = st.selectbox("Rola", ["użytkownik", "admin"], index=0 if p.get('rola') == 'użytkownik' else 1, key=f"er_{p['login']}")
                    if st.button("💾 Zapisz", key=f"es_{p['login']}", type="primary", use_container_width=True):
                        supabase.table("pracownicy").update({"haslo": e_has, "telefon": e_tel, "rola": e_rol}).eq("login", p['login']).execute()
                        st.toast(f"Zapisano zmiany dla {p['login']}")
                        time.sleep(0.5); st.rerun()

                if p['login'].lower() != "emil":
                    if col_del.button("🗑️ Usuń", key=f"dp_{p['login']}", use_container_width=True):
                        supabase.table("pracownicy").delete().eq("login", p['login']).execute(); st.rerun()

    elif menu == "📊 Statystyki":
        st.title("📊 Statystyki i Analityka")
        res = supabase.table("zamowienia").select("*").execute()
        
        if res.data:
            df = pd.DataFrame(res.data)
            
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Wszystkich zamówień", len(df))
            m2.metric("Oczekujące", len(df[df['status'] == 'Oczekujące']))
            m3.metric("Zrealizowane", len(df[df['status'] == 'Zrealizowane']))
            
            df['data_zgloszenia'] = pd.to_datetime(df['data_zgloszenia'])
            df['data_realizacji'] = pd.to_datetime(df['data_realizacji'])
            df_zrealizowane = df[df['data_realizacji'].notna()]
            if not df_zrealizowane.empty:
                sredni_czas = (df_zrealizowane['data_realizacji'] - df_zrealizowane['data_zgloszenia']).dt.days.mean()
                m4.metric("Średni czas realizacji", f"{sredni_czas:.1f} dni")
            else:
                m4.metric("Średni czas realizacji", "---")

            st.divider()

            col_a, col_b = st.columns(2)
            with col_a:
                st.subheader("👤 Zamówienia wg Osób")
                st.bar_chart(df['zgloszone_przez'].value_counts())
                st.subheader("🚨 Rozkład Pilności")
                st.bar_chart(df['pilnosc'].value_counts())
            with col_b:
                st.subheader("🏗️ Zamówienia wg Projektu")
                st.bar_chart(df['projekt'].value_counts())
                st.subheader("🔝 Top 5 Produktów")
                top_produkty = df['pozycja'].value_counts().head(5)
                st.table(top_produkty)
        else:
            st.info("Brak danych do wygenerowania statystyk.")

    elif menu == "💬 Czat / Sugestie":
        st.title("💬 Czat i Sugestie")
        st.caption("Masz pomysł na nową funkcję? Coś nie działa? Napisz tutaj!")
        
        with st.container(border=True):
            sugestia_tekst = st.text_area("Twoja sugestia / pytanie:", placeholder="Napisz tutaj co możemy ulepszyć...")
            if st.button("Wyślij wpis", type="primary"):
                if sugestia_tekst:
                    try:
                        supabase.table("sugestie").insert({
                            "data": datetime.now().strftime("%d.%m.%Y %H:%M"),
                            "uzytkownik": st.session_state.uzytkownik,
                            "tresc": sugestia_tekst,
                            "status": "Oczekujące"
                        }).execute()
                        st.success("Dziękujemy za sugestię!")
                        time.sleep(1); st.rerun()
                    except Exception as e:
                        st.error(f"Wystąpił błąd: Sprawdź uprawnienia RLS w Supabase.")
                else:
                    st.error("Wpisz treść!")

        st.divider()
        
        try:
            sug_res = supabase.table("sugestie").select("*").order("id", desc=True).execute()
            if not sug_res.data:
                st.info("Brak wpisów. Bądź pierwszy!")
            else:
                for s in sug_res.data:
                    with st.container(border=True):
                        c1, c2 = st.columns([4, 1])
                        c1.markdown(f"👤 **{s['uzytkownik']}** | 📅 `{s['data']}`")
                        s_color = "orange" if s['status'] == "Oczekujące" else "green"
                        c2.markdown(f"**Status:** :{s_color}[{s['status']}]")
                        st.markdown(f"> {s['tresc']}")
                        
                        if s.get('odpowiedz'):
                            st.markdown(f"**💬 Odpowiedź Admina:**")
                            st.info(s['odpowiedz'])
                        
                        if st.session_state.rola == "admin":
                            with st.expander("🛠️ Zarządzaj tym wpisem"):
                                nowa_odp = st.text_area("Twoja odpowiedź:", value=s.get('odpowiedz') or "", key=f"ans_{s['id']}")
                                nowy_stat = st.selectbox("Status sugestii:", ["Oczekujące", "Wprowadzone", "W trakcie", "Odrzucone"], 
                                                         index=["Oczekujące", "Wprowadzone", "W trakcie", "Odrzucone"].index(s['status']), 
                                                         key=f"stat_{s['id']}")
                                
                                col_a, col_b = st.columns(2)
                                if col_a.button("Zapisz odpowiedź", key=f"btn_s_{s['id']}", use_container_width=True, type="primary"):
                                    supabase.table("sugestie").update({"odpowiedz": nowa_odp, "status": nowy_stat}).eq("id", s['id']).execute()
                                    st.rerun()
                                if col_b.button("Usuń wpis", key=f"btn_d_{s['id']}", use_container_width=True):
                                    supabase.table("sugestie").delete().eq("id", s['id']).execute()
                                    st.rerun()
        except Exception as e:
            st.error(f"Nie udało się załadować tabeli sugestie: Upewnij się, że wykonałeś skrypt SQL do utworzenia tabeli.")

    elif menu == "📖 Instrukcja":
        st.title("📖 Instrukcja Obsługi Systemu")
        
        with st.container(border=True):
            st.markdown("### 📝 Składanie Zamówień")
            st.write("1. Wejdź w zakładkę **Nowe Zamówienie**.")
            st.write("2. Wypełnij nazwę pozycji, ilość oraz dane techniczne (wymiary, materiał).")
            st.write("3. Możesz zrobić zdjęcie telefonem (📷) lub wgrać plik PDF/Zdjęcie z dysku.")
            st.write("4. Wybierz Admina, którego chcesz powiadomić i kliknij **Wyślij**.")

        with st.container(border=True):
            st.markdown("### ⚙️ Zarządzanie Realizacją (Admin)")
            st.write("Wszystkie aktywne zamówienia są widoczne w **Panelu Realizacji**.")
            st.markdown("- **Zmiana statusu:** Wybierz nowy status z listy i wpisz notatkę, a następnie kliknij dyskietkę (**💾**).")
            st.markdown("- **Narzędzia:** Ikona zębatki (**⚙️**) pozwala na edycję szczegółów lub usunięcie zamówienia.")

        with st.container(border=True):
            st.markdown("### 🔎 Historia i Raporty")
            st.write("W sekcji **Historia i Raporty** możesz przeglądać wszystkie archiwalne zamówienia.")
            st.markdown("- **Excel:** Pełna tabela z kolorami, gotowa do dalszej obróbki.")
            st.markdown("- **PDF:** Przejrzysty dokument sformatowany pod wydruk **A4 w pionie**.")

        with st.container(border=True):
            st.markdown("### 💬 Czat i Sugestie")
            st.write("To miejsce komunikacji. Masz pomysł na ulepszenie systemu? Napisz tutaj. Admin odpowie Ci bezpośrednio pod Twoim wpisem.")
